from openpyxl import load_workbook
import pandas as pd
from langchain_groq import ChatGroq
from dotenv import load_dotenv
load_dotenv()

import os
api = os.getenv("GROQ_API_KEY")
llm = ChatGroq(model='llama-3.3-70b-versatile',api_key=api)


def Extract_And_Insight(path: str,
                        instruction: str,
                        header_index: int = 0,
                        columns_index: list[int] = [1, 3, 5],
                        chunk_size: int = 1000):

    # --------------------------
    # Excel Extractor (same code)
    # --------------------------

    wb = load_workbook(path)
    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    df = pd.DataFrame(data)

    df.columns = df.iloc[header_index]
    df = df.iloc[header_index+1:]
    df = df.iloc[:, columns_index]

    chunks = []
    for i in range(0, len(df), chunk_size):
        chunk = df.iloc[i:i+chunk_size]
        chunk.reset_index(drop=True, inplace=True)
        chunks.append(chunk)

    # --------------------------
    # LLM Insight (same code)
    # --------------------------

    final_output = []

    for idx, chunk in enumerate(chunks):

        Str_df = chunk.to_string(index=True)

        prompt = f"""
You are an Employee Reducer.

TASK RULES:
- You will receive two inputs: Data and Instruction.
- For each row in the Data, perform the action required by the Instruction.
- Add TWO new columns on the LEFT side:
    1. Action
    2. Reason
- Also add an Index column on the far left.
- Return ONLY the final data as comma-separated values (CSV).
- Do NOT add markdown, tables, borders, pipes, code fences, backticks, or explanations.
- Do NOT output ANY extra text — ONLY the CSV table.
- No comments, no notes, no narration.

CRITICAL CSV RULES (DO NOT BREAK):
1. If ANY field contains a comma, you MUST wrap it in double quotes.
2. All rows MUST have the same number of columns.
3. Do NOT use tab characters.
4. Do NOT insert spaces before or after commas.
5. Output EXACTLY valid CSV.

FORMAT REQUIREMENT:
The output MUST follow this structure:

Index,Action,Reason,<original column 1>,<original column 2>,...

EXAMPLE OF CORRECT CSV FORMAT:
Index,Action,Reason,JOB DESCRIPTION,Department
0,Keep,"Core operations required daily","Invoice Processing, Vendor Payments, Ledger Update","Accounts"
1,Fire,"Tasks can be automated","Customer Calls, Ticket Logging, Issue Escalation","Support"

IMPORTANT:
- Always wrap JOB DESCRIPTION and any other field in quotes if it contains a comma.
- The model must NEVER output unquoted comma-containing fields.

Data:
{Str_df}

Instruction:
{instruction}

"""

        result = llm.invoke(prompt).content
        final_output.append(result)

    return "\n".join(final_output)
