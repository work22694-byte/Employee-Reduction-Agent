import streamlit as st
import pandas as pd
import io
import re
from openpyxl import load_workbook
from Agent import Extract_And_Insight  # your no-chunk function

# ==========================================
# STREAMLIT CONFIG
# ==========================================
st.set_page_config(page_title="Employee Reducer", layout="wide")
st.title("Employee Reducer - Smart Column Selector")


# ==========================================
# UTILITY: Auto-deduplicate columns
# ==========================================
def deduplicate_columns(columns):
    seen = {}
    new_cols = []
    for col in columns:
        col = str(col).strip()
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}.{seen[col]}")
    return new_cols


# ==========================================
# ROBUST LLM TABLE PARSER
# ==========================================
def parse_llm_table(llm_output: str):
    """
    Detects CSV (with quoted fields), Markdown, or space tables
    inside LLM output and returns a clean DataFrame.
    """
    text = llm_output.strip()

    # -------------------------------
    # 1. CSV (comma-separated)
    # -------------------------------
    try:
        df = pd.read_csv(io.StringIO(text))
        return df
    except:
        pass

    # -------------------------------
    # 2. Markdown table
    # -------------------------------
    if text.startswith("|"):
        try:
            lines = [line.strip("|") for line in text.splitlines() if "|" in line]
            cleaned = "\n".join(lines)
            df = pd.read_csv(io.StringIO(cleaned), sep="|")
            df = df.dropna(axis=1, how='all')
            df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
            return df
        except:
            pass

    # -------------------------------
    # 3. Space-separated fallback
    # -------------------------------
    lines = text.split("\n")
    if len(lines) < 2:
        return None

    header = re.split(r'\s{2,}', lines[0].strip())
    rows = []
    for line in lines[1:]:
        parts = re.split(r'\s{2,}', line.strip())
        if len(parts) == len(header):
            rows.append(parts)

    if rows:
        return pd.DataFrame(rows, columns=header)

    return None


# ==========================================
# EXCEL UPLOAD
# ==========================================
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:

    # Read file bytes once
    file_bytes = uploaded_file.read()
    excel_preview = io.BytesIO(file_bytes)
    excel_for_llm = io.BytesIO(file_bytes)

    # Load workbook
    wb = load_workbook(excel_preview)
    ws = wb.active
    data = [list(row) for row in ws.iter_rows(values_only=True)]

    df = pd.DataFrame(data)
    df.columns = df.iloc[0].fillna("Unnamed")
    df = df.iloc[1:].reset_index(drop=True)

    # ===== AUTO-DEDUPE DUPLICATE COLUMNS =====
    df.columns = deduplicate_columns(df.columns)

    st.subheader("Preview of your Excel")
    st.dataframe(df.head(10), use_container_width=True)

    # ==========================================
    # COLUMN SELECTION
    # ==========================================
    st.subheader("Select Columns for Analysis")
    selected_cols = st.multiselect(
        "Select the columns you want to process",
        options=list(df.columns),
        default=list(df.columns)[:2]
    )

    # ==========================================
    # RUN LLM
    # ==========================================
    if selected_cols:

        column_indices = [df.columns.get_loc(c) for c in selected_cols]

        instruction = st.text_area("Instruction for LLM", "Apply employee reduction rules")

        if st.button("Run Analysis"):
            with st.spinner("Processing..."):

                llm_output = Extract_And_Insight(
                    path=excel_for_llm,
                    instruction=instruction,
                    header_index=0,
                    columns_index=column_indices
                )

            st.success("LLM Analysis Complete!")

            # ==========================================
            # TABLE PARSING
            # ==========================================
            df_result = parse_llm_table(llm_output)

            if df_result is not None:
                st.subheader("LLM Output as Table")
                st.dataframe(df_result, use_container_width=True)

                # Excel download
                out_buf = io.BytesIO()
                df_result.to_excel(out_buf, index=False)
                st.download_button(
                    "Download Output as Excel",
                    data=out_buf,
                    file_name="LLM_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            else:
                st.warning("Could not parse table. Showing raw output.")
                st.text_area("Raw Output", llm_output, height=400)
